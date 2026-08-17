"""
调用 evo-buss-oss channelonline/stat 接口，查询通道在线状态。
复用 video_client.py 中 EVOClient 的 token 获取逻辑。
"""
import csv
import json
import logging
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from video_client import EVOClient


requests.packages.urllib3.disable_warnings()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

REQUEST_TIMEOUT = 30
MAX_RETRIES = 3
RETRY_DELAY = 2


def _post_with_retry(session, url, headers, payload, label="请求", client=None):
    """带重试的 POST 请求，返回 (data, error_msg)。401时自动刷新 token 并重试。"""
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.post(url, headers=headers, data=json.dumps(payload),
                                timeout=REQUEST_TIMEOUT, verify=False)
            if resp.status_code == 200:
                try:
                    return resp.json(), None
                except json.JSONDecodeError:
                    last_error = f"响应非 JSON (状态码 200): {resp.text[:500]}"
                    logger.error(f"[{label}] {last_error}")
                    return None, last_error
            elif resp.status_code == 401:
                if client is not None and attempt == 1:
                    logger.warning(f"[{label}] Token 已过期 (401)，尝试刷新...")
                    if client._ensure_valid_token():
                        headers = {
                            'Authorization': client.token,
                            'Content-Type': 'application/json',
                        }
                        logger.info(f"[{label}] Token 刷新成功，重试请求")
                        continue
                last_error = f"Token 已过期 (401)"
                logger.error(f"[{label}] {last_error}")
                return None, last_error
            else:
                last_error = f"HTTP {resp.status_code}: {resp.text[:200]}"
                logger.warning(f"[{label}] 第 {attempt}/{MAX_RETRIES} 次失败: {last_error}")
        except requests.exceptions.Timeout:
            last_error = f"请求超时 (>{REQUEST_TIMEOUT}s)"
            logger.warning(f"[{label}] 第 {attempt}/{MAX_RETRIES} 次超时")
        except requests.exceptions.ConnectionError as e:
            last_error = f"连接失败: {e}"
            logger.warning(f"[{label}] 第 {attempt}/{MAX_RETRIES} 次连接错误")
        except Exception as e:
            last_error = f"未知错误: {type(e).__name__}: {e}"
            logger.error(f"[{label}] {last_error}")
            return None, last_error

        if attempt < MAX_RETRIES:
            time.sleep(RETRY_DELAY * attempt)

    logger.error(f"[{label}] 全部 {MAX_RETRIES} 次重试失败: {last_error}")
    return None, last_error


def _format_duration(duration) -> str:
    """将 duration 格式化为 '4天18小时55分钟27秒'；值为 0 的单位跳过。
    支持两种输入：
      1. dict: {'day': 4, 'hour': 18, 'minute': 55, 'second': 27}
      2. int/float/str: 总秒数，如 412527 -> '4天18小时55分钟27秒'
    """
    day = hour = minute = second = 0

    if isinstance(duration, dict):
        day = duration.get('day', 0) or 0
        hour = duration.get('hour', 0) or 0
        minute = duration.get('minute', 0) or 0
        second = duration.get('second', 0) or 0
    elif isinstance(duration, (int, float)):
        total = int(duration)
        day, rem = divmod(total, 86400)
        hour, rem = divmod(rem, 3600)
        minute, second = divmod(rem, 60)
    elif isinstance(duration, str):
        s = duration.strip()
        if s:
            try:
                total = int(float(s))
                day, rem = divmod(total, 86400)
                hour, rem = divmod(rem, 3600)
                minute, second = divmod(rem, 60)
            except ValueError:
                return ""
        else:
            return ""
    else:
        return ""

    parts = []
    if day:
        parts.append(f"{day}天")
    if hour:
        parts.append(f"{hour}小时")
    if minute:
        parts.append(f"{minute}分钟")
    if second:
        parts.append(f"{second}秒")
    return "".join(parts)


def _extract_total_count(result: dict) -> int:
    """从不同响应结构中提取 totalCount"""
    if not isinstance(result, dict):
        return 0

    if 'data' in result and isinstance(result['data'], dict):
        val = result['data'].get('totalCount')
        if val is not None:
            return int(val)

    if 'totalCount' in result:
        return int(result['totalCount'])

    if 'data' in result and isinstance(result['data'], dict):
        for key in ('totalCount', 'total', 'count'):
            val = result['data'].get(key)
            if val is not None:
                return int(val)

    return 0


def _get_field_value(item: dict, field_name: str) -> str:
    """从记录中提取字段值，兼容多种结构（递归查找）：
       - 顶层直接取值: item[field_name] 为基础类型 / {'value':'xxx'}
       - resourceInfo 容器: item['resourceInfo'][field_name] 及内部嵌套
       - fields / properties / extFields / ext 容器
    返回字符串，取不到返回空串。
    """
    if not isinstance(item, dict):
        return ""

    def _unwrap(val):
        """对 value/label/... 对象型字段解包；非字典直接返回字符串。"""
        if val is None:
            return ""
        if isinstance(val, dict):
            for sub in ('value', 'name', 'text', 'label', 'displayValue'):
                if sub in val and val[sub] is not None:
                    return str(val[sub])
            for v in val.values():
                if v is not None and v != "":
                    return str(v)
            return ""
        return str(val)

    # 1. 顶层字段
    if field_name in item:
        result = _unwrap(item[field_name])
        if result:
            return result

    # 2. 从容器中取：resourceInfo 优先（用户指定），然后 fields/properties/...
    containers = ('resourceInfo', 'fields', 'properties', 'extFields', 'ext')
    for container in containers:
        node = item.get(container)
        if not isinstance(node, dict):
            continue
        # 2a. 容器下直接是该字段
        if field_name in node:
            result = _unwrap(node[field_name])
            if result:
                return result
        # 2b. resourceInfo 下可能再嵌套 fields / properties
        if container == 'resourceInfo':
            for sub_container in ('fields', 'properties', 'extFields', 'ext'):
                sub_node = node.get(sub_container)
                if isinstance(sub_node, dict) and field_name in sub_node:
                    result = _unwrap(sub_node[field_name])
                    if result:
                        return result

    return ""


def _extract_company(ci016: str, org_code: str | list[str] = "") -> str:
    """从 CI016 路径中提取 company（公司名）。
    策略: 先找设备类型段（含'摄像机'或'记录仪'），取其前一段作为 company；
          找不到设备类型时，按 org_code 精确等于 "L03001001001" 取右数第6段，否则取右数第5段。
    例:
      '发电机/外雄水电站/瓯江水电/固定摄像机/生产期/水电厂站/国电平台' → 瓯江水电
      '日照莒县风电场/山东新能源/工作记录仪/生产期/新能源场站/国电平台' → 山东新能源
      '小城子风电场/优能康平风电/机组/光伏区摄像机/生产期/新能源场站/国电平台' → 优能康平风电
    """
    if not ci016:
        return ""
    parts = [p for p in ci016.split('/') if p != ""]

    # 1. 锚定设备类型段（含"摄像机"或"记录仪"），取其前一段作为 company
    #    若设备类型前一段是"机组"，则跳过取再前一段
    for i, p in enumerate(parts):
        if '摄像机' in p or '记录仪' in p:
            if i >= 2 and parts[i - 1] == '机组':
                return parts[i - 2]
            elif i >= 1:
                return parts[i - 1]
            break

    # 2. 回退: 按固定位置取
    idx = -6 if org_code == "L03001001001" else -5
    if len(parts) >= abs(idx):
        return parts[idx]
    return ""


def query_channel_count(client: EVOClient, org_id: str) -> dict | None:
    """调用 channel/channel_count 接口，根据 id 查询通道数量。"""
    url = (
        f"{client.base_url}/evo-apigw/evo-channel/{client.api_version}"
        "/device/channel/channel_count"
    )
    payload = {
        "id": org_id,
        "checkStat": 1,
        "checkNodes": [],
        "menu": 0,
        "type": "L03;00;1"
    }
    headers = {
        'Authorization': client.token,
        'Content-Type': 'application/json',
    }

    logger.info(f"请求 URL: {url}")
    logger.info(f"请求 payload: {json.dumps(payload, ensure_ascii=False)}")

    data, err = _post_with_retry(client.session, url, headers, payload, f"channel_count(id={org_id})", client=client)
    if err:
        return None
    return data


def query_channel_online_stat(client: EVOClient, org_code: str | list[str] = "") -> dict | None:
    """调用 channelonline/stat 接口"""
    url = (
        f"{client.base_url}/evo-apigw/evo-buss-oss/{client.api_version}"
        "/monitor/v1/channelonline/stat"
    )
    payload = {
        "orgCode": org_code,
        "containSub": True,
        "channelTypes": [1, 3],
        "isImportantPoint": False,
        "groupCode": ""
    }
    headers = {
        'Authorization': client.token,
        'Content-Type': 'application/json',
    }

    logger.info(f"请求 URL: {url}")
    logger.info(f"请求 payload: {json.dumps(payload, ensure_ascii=False)}")

    data, err = _post_with_retry(client.session, url, headers, payload, "stat", client=client)
    if err:
        return None
    return data


def query_channel_offline_total_count(client: EVOClient, org_code: str | list[str] = "") -> dict | None:
    """调用 channelonline/list/totalCount 接口"""
    url = (
        f"{client.base_url}/evo-apigw/evo-buss-oss/{client.api_version}"
        "/monitor/v1/channelonline/list/totalCount"
    )
    payload = {
        "condition": {
            "channelTypes": [1, 3],
            "orgCode": org_code,
            "containSub": True,
            "isImportantPoint": False,
            "groupCode": "",
            "keywordCiCodes": ["CI006", "CI010"],
            "statusDurationCompareType": 3,
            "statusDuration": "",
            "statusDurationCompareUnit": 0,
            "offlineDurationCompareType": None,
            "offlineDuration": "",
            "offlineDurationCompareUnit": 0,
            "offlineCountCompareType": None,
            "offlineCount": "",
            "deviceStatus": -1,
            "channelStatus": 0,
            "keyword": "",
            "searchFields": None
        },
        "page": 1,
        "pageSize": 20,
        "order": [{"key": "", "type": ""}],
        "fields": [
            "CI006", "CI016", "CI010", "CI004", "CI005", "CI031",
            "CI048", "CI051", "CI111", "CI017", "CI008", "CI011",
            "channelCreateTime", "deviceStatus", "channelStatus",
            "statusDuration", "statusChangeTime",
            "offlineDuration", "offlineCount"
        ]
    }
    headers = {
        'Authorization': client.token,
        'Content-Type': 'application/json',
    }

    logger.info(f"请求 URL: {url}")

    data, err = _post_with_retry(client.session, url, headers, payload, "totalCount", client=client)
    if err:
        return None
    return data


def query_channel_offline_list(client: EVOClient, page: int = 1, page_size: int = 512, org_code: str | list[str] = "") -> dict | None:
    """调用 channelonline/list 接口（离线明细列表）"""
    url = (
        f"{client.base_url}/evo-apigw/evo-buss-oss/{client.api_version}"
        "/monitor/v1/channelonline/list"
    )
    payload = {
        "condition": {
            "channelTypes": [1, 3],
            "orgCode": org_code,
            "containSub": True,
            "isImportantPoint": False,
            "groupCode": "",
            "keywordCiCodes": ["CI006", "CI010"],
            "statusDurationCompareType": 3,
            "statusDuration": "",
            "statusDurationCompareUnit": 0,
            "offlineDurationCompareType": None,
            "offlineDuration": "",
            "offlineDurationCompareUnit": 0,
            "offlineCountCompareType": None,
            "offlineCount": "",
            "deviceStatus": -1,
            "channelStatus": 0,
            "keyword": "",
            "reportStatus": 0,
            "searchFields": None
        },
        "page": page,
        "pageSize": page_size,
        "order": [{"key": "", "type": ""}],
        "fields": [
            "CI006", "CI016", "CI010", "CI004", "CI005", "CI031",
            "CI048", "CI051", "CI111", "CI017", "CI008", "CI011",
            "channelCreateTime", "deviceStatus", "channelStatus",
            "statusDuration", "statusChangeTime",
            "offlineDuration", "offlineCount"
        ]
    }
    headers = {
        'Authorization': client.token,
        'Content-Type': 'application/json',
    }

    data, err = _post_with_retry(client.session, url, headers, payload, f"list(page={page})", client=client)
    if err:
        return None
    return data


def _process_org(client: EVOClient, org_code: str | list[str],
                 detail_file: str, result_file: str) -> int:
    """处理单个组织编码的查询、去重、格式化与文件保存。返回 0 成功，1 失败。"""
    logger.info("=" * 60)
    logger.info(f"处理 orgCode={org_code}")
    logger.info(f"  明细文件: {detail_file}, 精简结果: {result_file}")
    logger.info("=" * 60)

    # 1. 查询 stat（汇总数据，仅打印）
    result = query_channel_online_stat(client, org_code=org_code)
    if result is None:
        logger.warning(f"[{org_code}] 查询 stat 失败，继续")
    else:
        print(f"\n[channelonline/stat 返回结果 for {org_code}]:")
        print(json.dumps(result, ensure_ascii=False, indent=2))

    # 2. 查询 totalCount（明细总数）
    result2 = query_channel_offline_total_count(client, org_code=org_code)
    if result2 is None:
        logger.error(f"[{org_code}] 查询 totalCount 失败")
        offlineTotal = 0
    else:
        print(f"\n[channelonline/list/totalCount 返回结果 for {org_code}]:")
        print(json.dumps(result2, ensure_ascii=False, indent=2))
        offlineTotal = _extract_total_count(result2)
        logger.info(f"[{org_code}] offlineTotalCount = {offlineTotal}")

    # 3. 分页查询 list（离线明细列表）
    PAGE_SIZE = 512
    all_results = []
    page = 1
    MAX_PAGES = 100
    logger.info(f"[{org_code}] 开始分页拉取: 每页 {PAGE_SIZE} 条")
    while page <= MAX_PAGES:
        logger.info(f"[{org_code}] 正在拉取第 {page} 页...")
        result3 = query_channel_offline_list(
            client, page=page, page_size=PAGE_SIZE, org_code=org_code)
        if result3 is None:
            logger.error(f"[{org_code}] 第 {page} 页查询失败，跳过")
            page += 1
            continue
        batch = result3.get('results', [])
        if not isinstance(batch, list):
            logger.error(f"[{org_code}] 第 {page} 页 results 格式异常: {type(batch).__name__}")
            break
        if not batch:
            logger.info(f"[{org_code}] 第 {page} 页返回空列表，拉取结束")
            break
        all_results.extend(batch)
        logger.info(f"[{org_code}] 第 {page} 页获取 {len(batch)} 条，累计 {len(all_results)}")

        if len(batch) < PAGE_SIZE:
            logger.info(f"[{org_code}] 第 {page} 页仅 {len(batch)} 条 (<{PAGE_SIZE})，拉取结束")
            break
        page += 1

    if page > MAX_PAGES:
        logger.warning(f"[{org_code}] 达到最大页数 {MAX_PAGES}，可能有数据未拉取")

    # 去重：按 channelCode 保留第一条
    seen = {}
    deduped = []
    dup_count = 0
    for item in all_results:
        if not isinstance(item, dict):
            continue
        code = item.get('channelCode', '')
        if code and code not in seen:
            seen[code] = item
            deduped.append(item)
        else:
            dup_count += 1
    if dup_count > 0:
        logger.info(f"[{org_code}] 去重: {len(all_results)} → {len(deduped)} (移除 {dup_count} 条重复)")

    # 提取 CI016、company、channelName、statusDuration
    simplified_new = []
    for item in deduped:
        ci016_val = _get_field_value(item, "CI016")
        simplified_new.append({
            "CI016": ci016_val,
            "company": _extract_company(ci016_val, org_code),
            "channelName": _get_field_value(item, "channelName"),
            "statusDuration": _format_duration(item.get("statusDuration"))
        })

    total_count = offlineTotal if offlineTotal else len(deduped)

    # 4a. 保存 detail 文件（完整原始数据）
    output_old = {
        'offlineTotalCount': total_count,
        'results': deduped
    }
    try:
        with open(detail_file, 'w', encoding='utf-8') as f:
            json.dump(output_old, f, ensure_ascii=False, indent=2)
        logger.info(f"[{org_code}] 已保存 {len(deduped)} 条完整记录到 {detail_file}")
    except OSError as e:
        logger.error(f"[{org_code}] 写入 {detail_file} 失败: {e}")
        return 1

    # 4b. 保存 result 文件（精简格式）
    output_new = {
        'offlineTotalCount': total_count,
        'results': simplified_new
    }
    try:
        with open(result_file, 'w', encoding='utf-8') as f:
            json.dump(output_new, f, ensure_ascii=False, indent=2)
        logger.info(f"[{org_code}] 已保存 {len(simplified_new)} 条精简记录到 {result_file}")
    except OSError as e:
        logger.error(f"[{org_code}] 写入 {result_file} 失败: {e}")
        return 1

    # 4c. 保存 CSV 文件（不含 CI016 字段）
    csv_file = result_file.rsplit('.', 1)[0] + '.csv'
    try:
        with open(csv_file, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['company', 'channelName', 'statusDuration'])
            for item in simplified_new:
                writer.writerow([
                    item.get('company', ''),
                    item.get('channelName', ''),
                    item.get('statusDuration', '')
                ])
        logger.info(f"[{org_code}] 已保存 {len(simplified_new)} 条记录到 {csv_file}")
    except OSError as e:
        logger.error(f"[{org_code}] 写入 {csv_file} 失败: {e}")
        return 1

    return 0


def _merge_ne_results():
    """合并 ne1/ne2/ne3 的 result JSON 和 CSV 为 ne_result.json / ne_result.csv。"""
    ne_sources = ['ne1_result_.json', 'ne2_result_.json', 'ne3_result_.json']
    all_results = []
    total_count = 0

    for src in ne_sources:
        try:
            with open(src, 'r', encoding='utf-8') as f:
                data = json.load(f)
            total_count += data.get('offlineTotalCount', 0)
            all_results.extend(data.get('results', []))
            logger.info(f"合并 {src}: {len(data.get('results', []))} 条")
        except FileNotFoundError:
            logger.warning(f"合并: {src} 不存在，跳过")
        except (json.JSONDecodeError, OSError) as e:
            logger.error(f"合并: 读取 {src} 失败: {e}")

    # 写入合并后的 ne_result.json
    merged = {
        'offlineTotalCount': total_count if total_count else len(all_results),
        'results': all_results
    }
    try:
        with open('ne_result.json', 'w', encoding='utf-8') as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)
        logger.info(f"已合并 {len(all_results)} 条记录到 ne_result.json")
    except OSError as e:
        logger.error(f"写入 ne_result.json 失败: {e}")

    # 写入合并后的 ne_result.csv
    try:
        with open('ne_result.csv', 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['channelName', 'site', 'company', 'statusDuration'])
            for item in all_results:
                ci016 = item.get('CI016', '')
                site = ci016.split('/')[0].strip() if ci016 else ''
                writer.writerow([
                    item.get('channelName', ''),
                    site,
                    item.get('company', ''),
                    item.get('statusDuration', '')
                ])
        logger.info(f"已合并 {len(all_results)} 条记录到 ne_result.csv")
    except OSError as e:
        logger.error(f"写入 ne_result.csv 失败: {e}")


def main():
    # ====== 可配置参数 ======
    # (org_code, detail_file, result_file) 列表，每个组织编码对应一对输出文件
    TASKS = [
        ("L03005001003",   "water_detail.json", "water_result_.json"),
        ("L03002001002",   "fire_detail.json",  "fire_result_.json"),
        ("L03001001001",   "ne1_detail.json",   "ne1_result_.json"),
        ("L03001001002",   "ne2_detail.json",   "ne2_result_.json"),
        ("L03001001003",   "ne3_detail.json",   "ne3_result_.json"),
    ]
    # ========================

    ids = [
        ("L03005001003",),
        ("L03002001002",),
        ("L03001001001", "L03001001002", "L03001001003"),
    ]
    try:
        with EVOClient(
            username="e0232091",
            password="Atos.202102",
            client_id="web_client",
            client_secret="web_client"
        ) as client:
            if not client._ensure_valid_token():
                logger.error("无法获取 Token，退出")
                return 1

            failed = 0
            for org_code, detail_file, result_file in TASKS:
                rc = _process_org(client, org_code, detail_file, result_file)
                if rc != 0:
                    failed += 1
                    logger.error(f"orgCode={org_code} 处理失败")

            # 合并 ne1/ne2/ne3 的 result JSON 和 CSV 为 ne_result
            if failed == 0:
                _merge_ne_results()

            if failed:
                logger.error(f"TASKS 有 {failed} 个 org 处理失败（继续执行 channel_count）")

            # 无论 TASKS 是否有失败，都调用 channel_count 接口并打印结果
            grand_total = 0
            grand_online = 0
            csv_rows = []
            group_names = ['水电合计', '火电合计', '新能源合计']
            for gi, group in enumerate(ids):
                label = "+".join(group)
                summary_name = group_names[gi] if gi < len(group_names) else f"group{gi}"
                logger.info(f"查询 channel_count, id={label}")
                print(f"\n[channel_count 精简结果 for {label}]:")
                sum_total = 0
                sum_online = 0
                for org_id in group:
                    result = query_channel_count(client, org_id)
                    if result is None:
                        logger.error(f"查询 channel_count 失败: id={org_id}")
                        continue
                    nodes = result.get('data', {}).get('value', [])
                    if not isinstance(nodes, list):
                        continue
                    for node in nodes:
                        total = node.get('total', 0) or 0
                        online = node.get('online', 0) or 0
                        if total == online:
                            continue
                        offline = total - online
                        orate_val = f"{offline / total * 100:.1f}" if total else "0"
                        orate = f"{orate_val}%"
                        print(f"  name={node.get('name', '')}  total={total}  offline={offline}  orate={orate}")
                        csv_rows.append([node.get('name', ''), total, offline, orate_val])
                        sum_total += total
                        sum_online += online
                sum_osum = sum_total - sum_online
                sum_orate_val = f"{sum_osum / sum_total * 100:.1f}" if sum_total else "0"
                sum_orate = f"{sum_orate_val}%"
                print(f"  [{summary_name}]  total={sum_total}  osum={sum_osum}  sumorate={sum_orate}")
                csv_rows.append([summary_name, sum_total, sum_osum, sum_orate_val])
                grand_total += sum_total
                grand_online += sum_online

            grand_osum = grand_total - grand_online
            grand_orate_val = f"{grand_osum / grand_total * 100:.1f}" if grand_total else "0"
            grand_orate = f"{grand_orate_val}%"
            print(f"\n[总计]  total={grand_total}  osum={grand_osum}  sumorate={grand_orate}")
            csv_rows.append(['总计', grand_total, grand_osum, grand_orate_val])

            # 保存到 summary.csv
            try:
                with open('summary.csv', 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['所属公司/组织', '设备通道总数', '离线通道总数', '设备离线率(%)'])
                    writer.writerows(csv_rows)
                logger.info("已保存 channel_count 统计到 summary.csv")
            except OSError as e:
                logger.error(f"写入 summary.csv 失败: {e}")

            # 保存到 summary.xlsx（格式化 Excel，多 sheet）
            try:
                wb = Workbook()

                # 样式定义
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_align = Alignment(horizontal="center", vertical="center")
                cell_align = Alignment(horizontal="center", vertical="center")
                name_align = Alignment(horizontal="left", vertical="center")
                summary_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                summary_font = Font(bold=True, size=11, color="FFFFFF")
                grand_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                grand_font = Font(bold=True, size=11, color="FFFFFF")
                alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                border = Border(
                    left=Side(style='thin', color='999999'),
                    right=Side(style='thin', color='999999'),
                    top=Side(style='thin', color='999999'),
                    bottom=Side(style='thin', color='999999')
                )

                # ===== Sheet 1: 各公司离线总数统计 =====
                ws1 = wb.active
                ws1.title = "各公司离线总数统计"
                ws1.freeze_panes = "A2"

                headers1 = ['所属公司/组织', '设备通道总数', '离线通道总数', '设备离线率(%)', '备注']
                for col, header in enumerate(headers1, 1):
                    cell = ws1.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = border

                for row_idx, row_data in enumerate(csv_rows, 2):
                    name = row_data[0]
                    is_summary = name in ('水电合计', '火电合计', '新能源合计', '总计')
                    for col, val in enumerate(row_data, 1):
                        cell = ws1.cell(row=row_idx, column=col, value=val)
                        cell.alignment = name_align if col == 1 else cell_align
                        cell.border = border
                        if is_summary:
                            cell.fill = grand_fill if name == '总计' else summary_fill
                            cell.font = grand_font if name == '总计' else summary_font
                        elif row_idx % 2 == 0:
                            cell.fill = alt_fill
                    # 备注列（第5列）
                    remark_cell = ws1.cell(row=row_idx, column=5, value='')
                    remark_cell.border = border
                    remark_cell.alignment = name_align
                    if is_summary:
                        remark_cell.fill = grand_fill if name == '总计' else summary_fill
                        remark_cell.font = grand_font if name == '总计' else summary_font
                    elif row_idx % 2 == 0:
                        remark_cell.fill = alt_fill

                ws1.column_dimensions['A'].width = 28
                ws1.column_dimensions['B'].width = 14
                ws1.column_dimensions['C'].width = 14
                ws1.column_dimensions['D'].width = 14
                ws1.column_dimensions['E'].width = 20
                ws1.row_dimensions[1].height = 22
                for _r in range(2, len(csv_rows) + 2):
                    ws1.row_dimensions[_r].height = 23
                ws1.auto_filter.ref = f"A1:E{len(csv_rows) + 1}"

                # ===== Sheet 2: 水电离线明细 =====
                ws2 = wb.create_sheet("水电离线明细")
                ws2.freeze_panes = "A2"

                headers2 = ['通道名称', '所属公司/组织', '当前状态持续时长', '备注']
                for col, header in enumerate(headers2, 1):
                    cell = ws2.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = border

                water_rows = 0
                try:
                    with open('water_result_.csv', 'r', encoding='utf-8-sig', newline='') as f:
                        reader = csv.reader(f)
                        next(reader, None)  # skip header
                        for row_idx, row in enumerate(reader, 2):
                            # csv 列顺序: company, channelName, statusDuration → 调整为 channelName, company, statusDuration
                            values = [
                                row[1] if len(row) > 1 else '',   # channelName → 第1列
                                row[0] if len(row) > 0 else '',   # company → 第2列
                                row[2] if len(row) > 2 else '',    # statusDuration → 第3列
                            ]
                            for col, val in enumerate(values, 1):
                                cell = ws2.cell(row=row_idx, column=col, value=val)
                                cell.alignment = name_align if col == 2 else cell_align
                                cell.border = border
                                if row_idx % 2 == 0:
                                    cell.fill = alt_fill
                            # 备注列（第4列）
                            remark_cell = ws2.cell(row=row_idx, column=4, value='')
                            remark_cell.border = border
                            remark_cell.alignment = name_align
                            if row_idx % 2 == 0:
                                remark_cell.fill = alt_fill
                            water_rows += 1
                except FileNotFoundError:
                    logger.warning("water_result_.csv 不存在，水电离线明细 sheet 为空")
                except OSError as e:
                    logger.error(f"读取 water_result_.csv 失败: {e}")

                ws2.column_dimensions['A'].width = 40
                ws2.column_dimensions['B'].width = 18
                ws2.column_dimensions['C'].width = 22
                ws2.column_dimensions['D'].width = 20
                ws2.row_dimensions[1].height = 22
                for _r in range(2, water_rows + 2):
                    ws2.row_dimensions[_r].height = 23
                ws2.auto_filter.ref = f"A1:D{water_rows + 1}"

                # ===== Sheet 3: 火电离线明细 =====
                ws3 = wb.create_sheet("火电离线明细")
                ws3.freeze_panes = "A2"

                headers3 = ['通道名称', '所属公司/组织', '当前状态持续时长', '备注']
                for col, header in enumerate(headers3, 1):
                    cell = ws3.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = border

                fire_rows = 0
                try:
                    with open('fire_result_.csv', 'r', encoding='utf-8-sig', newline='') as f:
                        reader = csv.reader(f)
                        next(reader, None)  # skip header
                        for row_idx, row in enumerate(reader, 2):
                            values = [
                                row[1] if len(row) > 1 else '',   # channelName → 第1列
                                row[0] if len(row) > 0 else '',   # company → 第2列
                                row[2] if len(row) > 2 else '',    # statusDuration → 第3列
                            ]
                            for col, val in enumerate(values, 1):
                                cell = ws3.cell(row=row_idx, column=col, value=val)
                                cell.alignment = name_align if col == 2 else cell_align
                                cell.border = border
                                if row_idx % 2 == 0:
                                    cell.fill = alt_fill
                            # 备注列（第4列）
                            remark_cell = ws3.cell(row=row_idx, column=4, value='')
                            remark_cell.border = border
                            remark_cell.alignment = name_align
                            if row_idx % 2 == 0:
                                remark_cell.fill = alt_fill
                            fire_rows += 1
                except FileNotFoundError:
                    logger.warning("fire_result_.csv 不存在，火电离线明细 sheet 为空")
                except OSError as e:
                    logger.error(f"读取 fire_result_.csv 失败: {e}")

                ws3.column_dimensions['A'].width = 40
                ws3.column_dimensions['B'].width = 18
                ws3.column_dimensions['C'].width = 22
                ws3.column_dimensions['D'].width = 20
                ws3.row_dimensions[1].height = 22
                for _r in range(2, fire_rows + 2):
                    ws3.row_dimensions[_r].height = 23
                ws3.auto_filter.ref = f"A1:D{fire_rows + 1}"

                # ===== Sheet 4: 新能源离线明细 =====
                ws4 = wb.create_sheet("新能源离线明细")
                ws4.freeze_panes = "A2"

                headers4 = ['通道名称', '场站名称', '所属公司/组织', '当前状态持续时长', '备注']
                for col, header in enumerate(headers4, 1):
                    cell = ws4.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = border

                ne_rows = 0
                try:
                    with open('ne_result.csv', 'r', encoding='utf-8-sig', newline='') as f:
                        reader = csv.reader(f)
                        next(reader, None)  # skip header
                        for row_idx, row in enumerate(reader, 2):
                            # CSV 列顺序: channelName, site, company, statusDuration → 直接对应
                            for col, val in enumerate(row, 1):
                                cell = ws4.cell(row=row_idx, column=col, value=val)
                                cell.alignment = name_align if col in (1, 2) else cell_align
                                cell.border = border
                                if row_idx % 2 == 0:
                                    cell.fill = alt_fill
                            # 备注列（第5列）
                            remark_cell = ws4.cell(row=row_idx, column=5, value='')
                            remark_cell.border = border
                            remark_cell.alignment = name_align
                            if row_idx % 2 == 0:
                                remark_cell.fill = alt_fill
                            ne_rows += 1
                except FileNotFoundError:
                    logger.warning("ne_result.csv 不存在，新能源离线明细 sheet 为空")
                except OSError as e:
                    logger.error(f"读取 ne_result.csv 失败: {e}")

                ws4.column_dimensions['A'].width = 40
                ws4.column_dimensions['B'].width = 20
                ws4.column_dimensions['C'].width = 18
                ws4.column_dimensions['D'].width = 22
                ws4.column_dimensions['E'].width = 20
                ws4.row_dimensions[1].height = 22
                for _r in range(2, ne_rows + 2):
                    ws4.row_dimensions[_r].height = 23
                ws4.auto_filter.ref = f"A1:E{ne_rows + 1}"

                wb.save('summary.xlsx')
                logger.info(f"已保存 summary.xlsx（各公司离线总数统计 + 水电 {water_rows} 条 + 火电 {fire_rows} 条 + 新能源 {ne_rows} 条）")
            except Exception as e:
                logger.error(f"写入 summary.xlsx 失败: {e}")

            return 0

    except KeyboardInterrupt:
        logger.info("用户中断")
        return 130
    except Exception as e:
        logger.error(f"程序异常: {type(e).__name__}: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    main()
